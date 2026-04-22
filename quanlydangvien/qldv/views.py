import csv
import datetime
import io
import os
import statistics
import unicodedata
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models.deletion import ProtectedError
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import (
    ChiBoForm,
    ChiBoImportForm,
    DangVienForm,
    DangVienImportForm,
    HuyHieuEditForm,
    HuyHieuImportForm,
    HuyHieuLegacyForm,
    ETHNIC_GROUPS,
    PROVINCE_CITY_CHOICES,
)
from .models import ChiBo, DangBo, DangVien, HuyHieuDang
from .services.huyhieu_service import (
    BADGE_MILESTONES,
    get_ceremony_dates,
    get_eligible_members,
    validate_badge_decision_number,
)


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value):
    raw = _normalize_text(value)
    normalized = unicodedata.normalize("NFD", raw)
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return normalized.lower().replace(" ", "").replace("_", "")


def _parse_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value

    text = _normalize_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_rows_from_csv(uploaded_file, header_aliases):
    raw = uploaded_file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    rows = []
    for row in reader:
        parsed = {}
        for key, value in row.items():
            canonical = header_aliases.get(_normalize_header(key), _normalize_text(key))
            parsed[canonical] = value
        rows.append(parsed)
    return rows


def _parse_rows_from_excel(uploaded_file, header_aliases):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Cần cài đặt gói openpyxl để import file Excel (.xlsx).") from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    headers = None

    for row_values in sheet.iter_rows(values_only=True):
        if headers is None:
            headers = [
                header_aliases.get(_normalize_header(header), _normalize_text(header))
                for header in row_values
            ]
            continue

        if not any(value is not None and str(value).strip() for value in row_values):
            continue

        parsed = {}
        for idx, value in enumerate(row_values):
            if idx >= len(headers):
                continue
            parsed[headers[idx]] = value
        rows.append(parsed)

    return rows


def _parse_uploaded_rows(uploaded_file, header_aliases):
    file_name = uploaded_file.name.lower()
    if file_name.endswith(".csv"):
        return _parse_rows_from_csv(uploaded_file, header_aliases)
    if file_name.endswith(".xlsx"):
        return _parse_rows_from_excel(uploaded_file, header_aliases)
    raise RuntimeError("Định dạng file không được hỗ trợ. Vui lòng dùng .csv hoặc .xlsx.")


def _get_single_dang_bo():
    default_dang_bo_name = "Đảng bộ Phường Sơn Trà"
    dang_bo = DangBo.objects.order_by("DangBoID").first()
    if dang_bo:
        if dang_bo.TenDangBo == "Đảng bộ mặc định":
            dang_bo.TenDangBo = default_dang_bo_name
            dang_bo.save(update_fields=["TenDangBo"])
        return dang_bo

    return DangBo.objects.create(
        TenDangBo=default_dang_bo_name,
        CapDangBo="Cấp cơ sở",
        TrangThai="Đang hoạt động",
    )


def _import_chi_bo_data(uploaded_file, default_dang_bo=None):
    header_aliases = {
        "tenchibo": "TenChiBo",
        "diaban": "DiaBan",
        "trangthai": "TrangThai",
        "dangboid": "DangBoID",
        "tendangbo": "TenDangBo",
    }
    rows = _parse_uploaded_rows(uploaded_file, header_aliases)

    created = 0
    errors = []
    dang_bo_by_id = {str(obj.pk): obj for obj in DangBo.objects.all()}
    dang_bo_by_name = {obj.TenDangBo.lower(): obj for obj in DangBo.objects.all()}
    existing_chi_bo_by_key = {
        ChiBo.normalized_name_key(obj.TenChiBo): obj.TenChiBo
        for obj in ChiBo.objects.all()
    }

    valid_dia_ban = ["Nại Hiên Đông", "Mân Thái", "Thọ Quang"]

    with transaction.atomic():
        for index, row in enumerate(rows, start=2):
            ten_chi_bo = _normalize_text(row.get("TenChiBo"))
            dia_ban = _normalize_text(row.get("DiaBan"))
            trang_thai = _normalize_text(row.get("TrangThai"))

            if not ten_chi_bo or not dia_ban or not trang_thai:
                errors.append(f"Dòng {index}: thiếu TenChiBo, DiaBan hoặc TrangThai.")
                continue

            ten_chi_bo_key = ChiBo.normalized_name_key(ten_chi_bo)
            if ten_chi_bo_key in existing_chi_bo_by_key:
                errors.append(
                    f"Dòng {index}: Chi bộ '{ten_chi_bo}' bị trùng với '{existing_chi_bo_by_key[ten_chi_bo_key]}'."
                )
                continue

            if dia_ban not in valid_dia_ban:
                errors.append(f"Dòng {index}: DiaBan '{dia_ban}' không hợp lệ. Phải là: {', '.join(valid_dia_ban)}")
                continue

            dang_bo = default_dang_bo
            if _normalize_text(row.get("DangBoID")):
                dang_bo = dang_bo_by_id.get(_normalize_text(row.get("DangBoID")))
            if not dang_bo and _normalize_text(row.get("TenDangBo")):
                dang_bo = dang_bo_by_name.get(_normalize_text(row.get("TenDangBo")).lower())
            if not dang_bo:
                errors.append(f"Dòng {index}: không tìm thấy Đảng bộ.")
                continue

            if default_dang_bo and dang_bo.pk != default_dang_bo.pk:
                errors.append(f"Dòng {index}: chỉ hỗ trợ 1 Đảng bộ trong hệ thống.")
                continue

            ChiBo.objects.create(
                TenChiBo=ten_chi_bo,
                DiaBan=dia_ban,
                TrangThai=trang_thai,
                DangBoID=dang_bo,
            )
            existing_chi_bo_by_key[ten_chi_bo_key] = ten_chi_bo
            created += 1

    return created, errors


def _import_dang_vien_data(uploaded_file, default_dang_bo=None):
    header_aliases = {
        "socccd": "SoCCCD",
        "cccd": "SoCCCD",
        "hodemvaten": "HoTen",
        "hoten": "HoTen",
        "gioitinh": "GioiTinh",
        "ngaysinh": "NgaySinh",
        "ngayvaodang": "NgayVaoDang",
        "ngaychinhthuc": "NgayChinhThuc",
        "dangboid": "DangBoID",
        "dangbo": "TenDangBo",
        "tendangbo": "TenDangBo",
        "chiboid": "ChiBoID",
        "chibo": "TenChiBo",
        "tenchibo": "TenChiBo",
        "trangthai": "TrangThaiSinhHoat",
        "trangthaisinhhoat": "TrangThaiSinhHoat",
        "bidanh": "BiDanh",
        "quequan": "QueQuan",
        "noicutru": "NoiThuongTru",
        "noithuongtru": "NoiThuongTru",
        "noitamtru": "NoiTamTru",
        "dantoc": "DanToc",
        "nghenghiep": "NgheNghiep",
        "gdpt": "GDPT",
        "gdnn": "GDNN",
        "gddh": "GDDH",
        "gdsđh": "GDSĐH",
        "gddhsdh": "GDDH",
        "hocham": "HocHam",
        "lyluanchinhtri": "LyLuanChinhTri",
        "ngoaingu": "NgoaiNgu",
        "tinhoc": "TinHoc",
        "tiengdtts": "TiengDTTS",
        "diendangvien": "DienDangVien",
        "sodienthoai": "SoDienThoai",
        "dienthoai": "SoDienThoai",
        "ghichu": "GhiChu",
        "huyhieucaonhat": "HuyHieuCaoNhat",
        "huyhieu": "HuyHieuCaoNhat",
        "sohuyhieu": "SoHuyHieu",
    }
    rows = _parse_uploaded_rows(uploaded_file, header_aliases)

    required_columns = [
        "SoCCCD",
        "HoTen",
        "GioiTinh",
        "NgaySinh",
        "NgayVaoDang",
        "DienDangVien",
        "TrangThaiSinhHoat",
    ]
    created = 0
    errors = []
    valid_trang_thai_sinh_hoat = {
        choice[0] for choice in DangVien.TRANG_THAI_SINH_HOAT_CHOICES
    }

    dang_bo_by_id = {str(obj.pk): obj for obj in DangBo.objects.all()}
    dang_bo_by_name = {obj.TenDangBo.lower(): obj for obj in DangBo.objects.all()}
    chi_bo_by_id = {str(obj.pk): obj for obj in ChiBo.objects.select_related("DangBoID")}
    chi_bo_by_name = {obj.TenChiBo.lower(): obj for obj in ChiBo.objects.select_related("DangBoID")}

    with transaction.atomic():
        for index, row in enumerate(rows, start=2):
            missing = [column for column in required_columns if not _normalize_text(row.get(column))]
            if missing:
                errors.append(f"Dòng {index}: thiếu cột bắt buộc: {', '.join(missing)}")
                continue

            so_cccd = _normalize_text(row.get("SoCCCD"))

            if not so_cccd.isdigit() or len(so_cccd) != 12:
                errors.append(f"Dòng {index}: Số CCCD phải là 12 chữ số ('{so_cccd}').")
                continue

            if DangVien.objects.filter(SoCCCD=so_cccd).exists():
                errors.append(f"Dòng {index}: Số CCCD '{so_cccd}' đã tồn tại.")
                continue

            gioi_tinh_raw = _normalize_text(row.get("GioiTinh")).lower()
            if gioi_tinh_raw in {"nam", "male"}:
                gioi_tinh = "Nam"
            elif gioi_tinh_raw in {"nu", "nữ", "female"}:
                gioi_tinh = "Nữ"
            else:
                errors.append(f"Dòng {index}: GioiTinh không hợp lệ ('{row.get('GioiTinh')}').")
                continue

            ngay_sinh = _parse_date(row.get("NgaySinh"))
            ngay_vao_dang = _parse_date(row.get("NgayVaoDang"))
            dien_dang_vien = DangVien._normalize_membership_value(_normalize_text(row.get("DienDangVien")))

            if not ngay_sinh or not ngay_vao_dang:
                errors.append(f"Dòng {index}: NgaySinh hoặc NgayVaoDang sai định dạng.")
                continue

            dang_bo = default_dang_bo
            if _normalize_text(row.get("DangBoID")):
                dang_bo = dang_bo_by_id.get(_normalize_text(row.get("DangBoID")))
            if not dang_bo and _normalize_text(row.get("TenDangBo")):
                dang_bo = dang_bo_by_name.get(_normalize_text(row.get("TenDangBo")).lower())
            if not dang_bo:
                errors.append(f"Dòng {index}: không tìm thấy Đảng bộ.")
                continue

            if default_dang_bo and dang_bo.pk != default_dang_bo.pk:
                errors.append(f"Dòng {index}: chỉ hỗ trợ 1 Đảng bộ trong hệ thống.")
                continue

            chi_bo = None
            if _normalize_text(row.get("ChiBoID")):
                chi_bo = chi_bo_by_id.get(_normalize_text(row.get("ChiBoID")))
            if not chi_bo and _normalize_text(row.get("TenChiBo")):
                chi_bo = chi_bo_by_name.get(_normalize_text(row.get("TenChiBo")).lower())
            if not chi_bo:
                errors.append(f"Dòng {index}: không tìm thấy Chi bộ.")
                continue

            if chi_bo.DangBoID_id != dang_bo.pk:
                errors.append(f"Dòng {index}: Chi bộ không thuộc Đảng bộ đã chọn.")
                continue

            so_dien_thoai = _normalize_text(row.get("SoDienThoai"))
            if so_dien_thoai and (not so_dien_thoai.isdigit() or len(so_dien_thoai) != 10):
                errors.append(f"Dòng {index}: Số điện thoại phải là 10 chữ số ('{so_dien_thoai}').")
                continue

            trang_thai_sinh_hoat = _normalize_text(row.get("TrangThaiSinhHoat"))
            if trang_thai_sinh_hoat not in valid_trang_thai_sinh_hoat:
                errors.append(
                    f"Dòng {index}: Trạng thái sinh hoạt không hợp lệ ('{trang_thai_sinh_hoat}')."
                )
                continue

            DangVien.objects.create(
                SoCCCD=so_cccd,
                HoTen=_normalize_text(row.get("HoTen")),
                BiDanh=_normalize_text(row.get("BiDanh")) or None,
                GioiTinh=gioi_tinh,
                NgaySinh=ngay_sinh,
                QueQuan=_normalize_text(row.get("QueQuan")) or None,
                NoiThuongTru=_normalize_text(row.get("NoiThuongTru")) or None,
                NoiTamTru=_normalize_text(row.get("NoiTamTru")) or None,
                DanToc=_normalize_text(row.get("DanToc")) or None,
                NgheNghiep=_normalize_text(row.get("NgheNghiep")) or None,
                GDPT=_normalize_text(row.get("GDPT")) or None,
                GDNN=_normalize_text(row.get("GDNN")) or None,
                GDDH=_normalize_text(row.get("GDDH")) or None,
                GDSĐH=_normalize_text(row.get("GDSĐH")) or None,
                HocHam=_normalize_text(row.get("HocHam")) or None,
                LyLuanChinhTri=_normalize_text(row.get("LyLuanChinhTri")) or None,
                NgoaiNgu=_normalize_text(row.get("NgoaiNgu")) or None,
                TinHoc=_normalize_text(row.get("TinHoc")) or None,
                TiengDTTS=_normalize_text(row.get("TiengDTTS")) or None,
                NgayVaoDang=ngay_vao_dang,
                ChiBoID=chi_bo,
                DangBoID=dang_bo,
                TrangThaiSinhHoat=trang_thai_sinh_hoat,
                DienDangVien=dien_dang_vien,
                SoDienThoai=so_dien_thoai or None,
                GhiChu=_normalize_text(row.get("GhiChu")) or None,
                HuyHieuCaoNhat=_normalize_text(row.get("HuyHieuCaoNhat")) or None,
                SoHuyHieu=int(_normalize_text(row.get("SoHuyHieu")) or 0),
            )
            created += 1

    return created, errors


def _resolve_dang_vien_for_huy_hieu(ma_dang_vien=None, so_cccd=None, dang_vien_id=None):
    if dang_vien_id:
        try:
            return DangVien.objects.get(pk=int(str(dang_vien_id).strip()))
        except (DangVien.DoesNotExist, ValueError, TypeError):
            return None

    if ma_dang_vien:
        ma = str(ma_dang_vien).strip()
        if ma:
            return DangVien.objects.filter(MaDangVien=ma).first()

    if so_cccd:
        cccd = str(so_cccd).strip()
        if cccd:
            return DangVien.objects.filter(SoCCCD=cccd).first()

    return None


def _import_huy_hieu_data(uploaded_file):
    header_aliases = {
        "dangvienid": "DangVienID",
        "madangvien": "MaDangVien",
        "socccd": "SoCCCD",
        "loaihuyhieu": "LoaiHuyHieu",
        "ngaydudieukien": "NgayDuDieuKien",
        "dottraotang": "DotTraoTang",
        "trangthai": "TrangThai",
        "soquyetdinh": "SoQuyetDinh",
        "ngaytrao": "NgayTrao",
        "ghichu": "GhiChu",
    }
    rows = _parse_uploaded_rows(uploaded_file, header_aliases)

    created = 0
    errors = []

    with transaction.atomic():
        for index, row in enumerate(rows, start=2):
            dang_vien = _resolve_dang_vien_for_huy_hieu(
                ma_dang_vien=_normalize_text(row.get("MaDangVien")),
                so_cccd=_normalize_text(row.get("SoCCCD")),
                dang_vien_id=_normalize_text(row.get("DangVienID")),
            )
            if not dang_vien:
                errors.append(f"Dòng {index}: không tìm thấy Đảng viên theo DangVienID/MaDangVien/SoCCCD.")
                continue

            loai_huy_hieu = _normalize_text(row.get("LoaiHuyHieu"))
            dot_trao_tang = _normalize_text(row.get("DotTraoTang"))
            trang_thai = _normalize_text(row.get("TrangThai"))
            so_quyet_dinh = _normalize_text(row.get("SoQuyetDinh"))
            ngay_du_dieu_kien = _parse_date(row.get("NgayDuDieuKien"))
            ngay_trao = _parse_date(row.get("NgayTrao"))

            if not all([loai_huy_hieu, dot_trao_tang, trang_thai, so_quyet_dinh, ngay_du_dieu_kien, ngay_trao]):
                errors.append(f"Dòng {index}: thiếu dữ liệu bắt buộc hoặc sai định dạng ngày.")
                continue

            duplicated = HuyHieuDang.objects.filter(
                DangVienID=dang_vien,
                LoaiHuyHieu=loai_huy_hieu,
                SoQuyetDinh=so_quyet_dinh,
            ).exists()
            if duplicated:
                errors.append(f"Dòng {index}: hồ sơ đã tồn tại ({dang_vien.HoTen} - {so_quyet_dinh}).")
                continue

            HuyHieuDang.objects.create(
                DangVienID=dang_vien,
                LoaiHuyHieu=loai_huy_hieu,
                NgayDuDieuKien=ngay_du_dieu_kien,
                DotTraoTang=dot_trao_tang,
                TrangThai=trang_thai,
                SoQuyetDinh=so_quyet_dinh,
                NgayTrao=ngay_trao,
                GhiChu=_normalize_text(row.get("GhiChu")) or None,
            )
            created += 1

    return created, errors


def home(request):
    stats = {
        "chi_bo": ChiBo.objects.count(),
        "dang_vien": DangVien.objects.count(),
        "huy_hieu": HuyHieuDang.objects.count(),
    }

    quick_actions = [
        {
            "title": "Quản lý Chi bộ",
            "description": "Cập nhật thông tin Chi bộ theo từng khu vực.",
            "path": "/chibo/",
            "badge": "Đơn vị",
        },
        {
            "title": "Danh sách Đảng viên",
            "description": "Tra cứu hồ sơ, quá trình công tác và trạng thái sinh hoạt.",
            "path": "/dangvien/",
            "badge": "Nhân sự",
        },
        {
            "title": "Quản lý Huy hiệu Đảng",
            "description": "Tổng hợp đề nghị, phê duyệt và trao tặng huy hiệu.",
            "path": "/huyhieu/",
            "badge": "Khen thưởng",
        },
    ]

    recent_updates = ChiBo.objects.select_related("DangBoID").annotate(
        dang_vien=Count("dangvien", distinct=True)
    ).order_by("-ChiBoID")[:8]

    context = {
        "stats": stats,
        "quick_actions": quick_actions,
        "recent_updates": recent_updates,
    }
    return render(request, "qldv/home.html", context)


def _base_chi_bo_queryset():
    return ChiBo.objects.select_related("DangBoID").annotate(
        so_dang_vien=Count("dangvien", distinct=True)
    ).order_by("TenChiBo")


def _classify_chi_bo_category(name):
    text = (name or "").strip().lower()
    if any(k in text for k in ["trường", "thcs", "tiểu học", "mầm non", "thpt", "hoc"]):
        return "school"
    if any(k in text for k in ["cơ quan", "ubnd", "công an", "y tế", "đơn vị", "chi cục", "ban "]):
        return "agency"
    return "residential"


def _apply_chi_bo_filters(queryset, query_params):
    q = (query_params.get("q") or "").strip()
    dia_ban = (query_params.get("dia_ban") or "").strip()
    trang_thai = (query_params.get("trang_thai") or "").strip()

    if q:
        queryset = queryset.filter(TenChiBo__icontains=q)

    if dia_ban:
        queryset = queryset.filter(DiaBan=dia_ban)

    if trang_thai:
        queryset = queryset.filter(TrangThai=trang_thai)

    return queryset, {
        "q": q,
        "dia_ban": dia_ban,
        "trang_thai": trang_thai,
    }


def _export_chi_bo_csv(chibo_list):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="chi_bo_data.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow([
        "Tên Chi bộ",
        "Đảng bộ",
        "Địa bàn",
        "Trạng thái",
        "Số Đảng viên",
    ])

    for item in chibo_list:
        writer.writerow([
            item.TenChiBo,
            item.DangBoID.TenDangBo if item.DangBoID else "",
            item.DiaBan,
            item.TrangThai,
            item.so_dang_vien,
        ])

    return response


def _export_chi_bo_excel(chibo_list):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("Cần cài đặt gói openpyxl để xuất file Excel.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ChiBo"
    worksheet.append([
        "Tên Chi bộ",
        "Đảng bộ",
        "Địa bàn",
        "Trạng thái",
        "Số Đảng viên",
    ])

    for item in chibo_list:
        worksheet.append([
            item.TenChiBo,
            item.DangBoID.TenDangBo if item.DangBoID else "",
            item.DiaBan,
            item.TrangThai,
            item.so_dang_vien,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="chi_bo_data.xlsx"'
    workbook.save(response)
    return response


def chibo(request):
    # ====================== LẤY DỮ LIỆU CƠ BẢN ======================
    source_queryset = _base_chi_bo_queryset()
    source_rows = [
        {
            "ChiBoID": item.ChiBoID,
            "TenChiBo": item.TenChiBo,
            "DiaBan": item.DiaBan,
            "TrangThai": item.TrangThai,
            "so_dang_vien": item.so_dang_vien,
        }
        for item in source_queryset
    ]

    total_chibo_all = len(source_rows)

    # ====================== XỬ LÝ BỘ LỌC ======================
    selected_dia_ban = (request.GET.get("dia_ban") or "").strip()
    selected_trang_thai = (request.GET.get("trang_thai") or "").strip()
    selected_loai = (request.GET.get("loai") or "").strip()   # residential, agency, school

    filtered_rows = source_rows[:]

    if selected_dia_ban:
        filtered_rows = [row for row in filtered_rows if row["DiaBan"] == selected_dia_ban]

    if selected_trang_thai:
        filtered_rows = [
            row for row in filtered_rows 
            if _normalize_text(row["TrangThai"]).lower() == _normalize_text(selected_trang_thai).lower()
        ]

    if selected_loai:
        def classify(name):
            text = (name or "").strip().lower()
            if any(k in text for k in ["trường", "thcs", "tiểu học", "mầm non", "thpt", "hoc"]):
                return "school"
            if any(k in text for k in ["cơ quan", "ubnd", "công an", "y tế", "đơn vị", "chi cục", "ban "]):
                return "agency"
            return "residential"
        
        filtered_rows = [row for row in filtered_rows if classify(row["TenChiBo"]) == selected_loai]

    # ====================== TÍNH TOÁN THỐNG KÊ SAU KHI LỌC ======================
    total_chibo = len(filtered_rows)

    def _is_active(value):
        return _normalize_text(value).lower() in {"hoạt động", "đang hoạt động", "dang hoat dong", "hoat dong"}

    active_chibo = sum(1 for row in filtered_rows if _is_active(row["TrangThai"]))
    inactive_chibo = max(0, total_chibo - active_chibo)

    # Phân bổ theo địa bàn
    dia_ban_order = ["Thọ Quang", "Mân Thái", "Nại Hiên Đông"]
    dia_ban_colors = {
        "Thọ Quang": "#be123c",
        "Mân Thái": "#9f1239",
        "Nại Hiên Đông": "#b45309",
    }

    by_dia_ban = []
    for dia_ban in dia_ban_order:
        area_rows = [r for r in filtered_rows if r["DiaBan"] == dia_ban]
        count = len(area_rows)
        active_count = sum(1 for r in area_rows if _is_active(r["TrangThai"]))
        inactive_count = max(0, count - active_count)
        percent = round((count * 100 / total_chibo), 1) if total_chibo else 0

        by_dia_ban.append({
            "name": dia_ban,
            "count": count,
            "active_count": active_count,
            "inactive_count": inactive_count,
            "percent": percent,
            "color": dia_ban_colors.get(dia_ban, "#64748b"),
        })

    # Donut style
    start = 0.0
    donut_parts = []
    for item in by_dia_ban:
        if item["count"] > 0:
            end = start + item["percent"]
            donut_parts.append(f"{item['color']} {start:.4f}% {end:.4f}%")
            start = end
    if start < 100 and total_chibo > 0:
        donut_parts.append(f"#e2e8f0 {start:.4f}% 100%")
    dia_ban_donut_style = f"conic-gradient({', '.join(donut_parts)})" if donut_parts else "conic-gradient(#cbd5e1 0 100%)"

    # Bar ratio
    max_area_count = max((item["count"] for item in by_dia_ban), default=0)
    for item in by_dia_ban:
        item["bar_ratio"] = (item["count"] * 100 / max_area_count) if max_area_count else 0

    # Phân loại Chi bộ (residential, agency, school)
    category_map = {
        "residential": {"label": "Khu dân cư", "icon": "🏠", "color": "#be123c"},
        "agency": {"label": "Cơ quan / Đơn vị", "icon": "🏛️", "color": "#9f1239"},
        "school": {"label": "Trường học", "icon": "🏫", "color": "#b45309"},
    }

    category_counts = {"residential": 0, "agency": 0, "school": 0}
    category_names = {"residential": [], "agency": [], "school": []}

    for item in filtered_rows:
        key = _classify_chi_bo_category(item["TenChiBo"])   # hàm đã có sẵn trong code cũ
        category_counts[key] += 1
        category_names[key].append(item["TenChiBo"])

    category_blocks = []
    for key in ["residential", "agency", "school"]:
        cfg = category_map[key]
        count = category_counts[key]
        percent = round((count * 100 / total_chibo), 1) if total_chibo else 0
        sample = ", ".join(category_names[key][:3]) or "Chưa có dữ liệu"
        category_blocks.append({
            "label": cfg["label"],
            "icon": cfg["icon"],
            "count": count,
            "percent": percent,
            "color": cfg["color"],
            "description": sample,
        })

    residential_count = category_counts["residential"]
    active_percent = round((active_chibo * 100 / total_chibo), 1) if total_chibo else 0
    inactive_percent = round((inactive_chibo * 100 / total_chibo), 1) if total_chibo else 0
    residential_percent = round((residential_count * 100 / total_chibo), 1) if total_chibo else 0

    # ====================== CONTEXT ======================
    context = {
        "total_chibo": total_chibo,
        "active_chibo": active_chibo,
        "inactive_chibo": inactive_chibo,
        "residential_count": residential_count,
        "active_percent": active_percent,
        "inactive_percent": inactive_percent,
        "residential_percent": residential_percent,

        "by_dia_ban": by_dia_ban,
        "dia_ban_donut_style": dia_ban_donut_style,
        "category_blocks": category_blocks,

        # Dùng cho form lọc
        "selected_dia_ban": selected_dia_ban,
        "selected_trang_thai": selected_trang_thai,
        "selected_loai": selected_loai,
        "dia_ban_options": sorted({row["DiaBan"] for row in source_rows if row["DiaBan"]}),
        "trang_thai_options": sorted({row["TrangThai"] for row in source_rows if row["TrangThai"]}),
    }

    return render(request, "qldv/chibo.html", context)


def chibo_add(request):
    default_dang_bo = _get_single_dang_bo()

    if request.method == "POST" and request.POST.get("form_type") == "single":
        single_form = ChiBoForm(request.POST, dang_bo=default_dang_bo)
        import_form = ChiBoImportForm()
        if single_form.is_valid():
            single_form.save()
            messages.success(request, "Đã thêm Chi bộ thành công.")
            return redirect("chibo_add")
    elif request.method == "POST" and request.POST.get("form_type") == "import":
        single_form = ChiBoForm(dang_bo=default_dang_bo)
        import_form = ChiBoImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            try:
                created, errors = _import_chi_bo_data(
                    import_form.cleaned_data["import_file"],
                    default_dang_bo=default_dang_bo,
                )
            except Exception as exc:
                messages.error(request, f"Không thể import file: {exc}")
                return redirect("chibo_add")

            if created:
                messages.success(request, f"Đã import {created} Chi bộ.")
            if errors:
                messages.warning(request, "\n".join(errors[:20]))
            return redirect("chibo_add")
    else:
        single_form = ChiBoForm(dang_bo=default_dang_bo)
        import_form = ChiBoImportForm()

    return render(
        request,
        "qldv/chibo_add.html",
        {
            "single_form": single_form,
            "import_form": import_form,
        },
    )


def chibo_data(request):
    queryset = _base_chi_bo_queryset()
    filtered_queryset, active_filters = _apply_chi_bo_filters(queryset, request.GET)
    chibo_list = list(filtered_queryset)

    download_format = (request.GET.get("download") or "").strip().lower()
    if download_format == "csv":
        return _export_chi_bo_csv(chibo_list)
    if download_format == "excel":
        return _export_chi_bo_excel(chibo_list)

    dia_ban_options = (
        ChiBo.objects.exclude(DiaBan__isnull=True).exclude(DiaBan__exact="")
        .values_list("DiaBan", flat=True)
        .distinct()
    )
    trang_thai_options = (
        ChiBo.objects.exclude(TrangThai__isnull=True).exclude(TrangThai__exact="")
        .values_list("TrangThai", flat=True)
        .distinct()
    )

    return render(
        request,
        "qldv/chibo_data.html",
        {
            "chibo_list": chibo_list,
            "total_filtered": len(chibo_list),
            "filters": active_filters,
            "current_full_path": request.get_full_path(),
            "dia_ban_options": sorted(dia_ban_options),
            "trang_thai_options": sorted(trang_thai_options),
        },
    )


def chibo_edit(request, chibo_id):
    default_dang_bo = _get_single_dang_bo()
    chibo_item = get_object_or_404(ChiBo, pk=chibo_id)

    if request.method == "POST":
        form = ChiBoForm(request.POST, instance=chibo_item, dang_bo=default_dang_bo)
        if form.is_valid():
            form.save()
            messages.success(request, "Đã cập nhật Chi bộ.")
            return redirect("chibo_data")
    else:
        form = ChiBoForm(instance=chibo_item, dang_bo=default_dang_bo)

    return render(
        request,
        "qldv/chibo_edit.html",
        {
            "form": form,
            "chibo_item": chibo_item,
        },
    )


def chibo_delete(request, chibo_id):
    if request.method != "POST":
        return redirect("chibo_data")

    chibo_item = get_object_or_404(ChiBo, pk=chibo_id)
    try:
        chibo_item.delete()
    except ProtectedError:
        messages.error(
            request,
            "Không thể xóa Chi bộ vì đang có Đảng viên liên kết. Vui lòng chuyển hoặc xóa Đảng viên trước.",
        )
    else:
        messages.success(request, "Đã xóa Chi bộ thành công.")

    next_path = (request.POST.get("next") or "").strip()
    if next_path.startswith("/"):
        return redirect(next_path)
    return redirect("chibo_data")


def _sync_dang_vien_membership_statuses(reference_date=None):
    reference = reference_date or timezone.localdate()
    for member in DangVien.objects.all():
        old_status = member.DienDangVien
        old_official_date = member.NgayChinhThuc
        member.apply_membership_rules(reference_date=reference)
        if old_status != member.DienDangVien or old_official_date != member.NgayChinhThuc:
            member.save(update_fields=["DienDangVien", "NgayChinhThuc"])


def _base_dang_vien_queryset():
    return DangVien.objects.select_related("DangBoID", "ChiBoID").annotate(
        so_huy_hieu=Count("huyhieudang", distinct=True)
    ).order_by("HoTen")


def _apply_dang_vien_filters(queryset, query_params):
    keyword = (query_params.get("q") or "").strip()
    gioi_tinh = (query_params.get("gioi_tinh") or "").strip()
    dien = (query_params.get("dien") or "").strip()
    chi_bo = (query_params.get("chi_bo") or "").strip()
    trang_thai = (query_params.get("trang_thai") or "").strip()
    que_quan = (query_params.get("que_quan") or "").strip()
    dan_toc = (query_params.get("dan_toc") or "").strip()
    gdpt = (query_params.get("gdpt") or "").strip()
    gddh_group = (query_params.get("gddh_group") or "").strip()
    gdsdh_group = (query_params.get("gdsdh_group") or "").strip()
    hoc_ham_group = (query_params.get("hoc_ham_group") or "").strip()
    ngoai_ngu_group = (query_params.get("ngoai_ngu_group") or "").strip()
    tin_hoc_group = (query_params.get("tin_hoc_group") or "").strip()
    ly_luan_chinh_tri = (query_params.get("ly_luan_chinh_tri") or "").strip()
    tuoi_dang_nhom = (query_params.get("tuoi_dang_nhom") or "").strip()

    if keyword:
        queryset = queryset.filter(
            Q(HoTen__icontains=keyword)
            | Q(MaDangVien__icontains=keyword)
            | Q(SoCCCD__icontains=keyword)
            | Q(SoDienThoai__icontains=keyword)
        )

    if gioi_tinh:
        queryset = queryset.filter(GioiTinh=gioi_tinh)

    if trang_thai:
        queryset = queryset.filter(TrangThaiSinhHoat=trang_thai)

    if dien:
        queryset = queryset.filter(DienDangVien=dien)

    if chi_bo:
        queryset = queryset.filter(ChiBoID_id=chi_bo)

    if que_quan:
        queryset = queryset.filter(QueQuan__icontains=que_quan)

    if dan_toc:
        queryset = queryset.filter(DanToc__icontains=dan_toc)

    if gdpt:
        queryset = queryset.filter(GDPT__icontains=gdpt)

    if ly_luan_chinh_tri:
        queryset = queryset.filter(LyLuanChinhTri=ly_luan_chinh_tri)

    def apply_khong_khac_filter(current_queryset, field_name, value):
        if not value:
            return current_queryset
        if value == "KHONG":
            return current_queryset.filter(**{f"{field_name}__iexact": "Không"})
        if value == "KHAC":
            return current_queryset.exclude(**{f"{field_name}__isnull": True}).exclude(
                **{f"{field_name}__exact": ""}
            ).exclude(**{f"{field_name}__iexact": "Không"})
        return current_queryset

    queryset = apply_khong_khac_filter(queryset, "GDDH", gddh_group)
    queryset = apply_khong_khac_filter(queryset, "GDSĐH", gdsdh_group)
    queryset = apply_khong_khac_filter(queryset, "HocHam", hoc_ham_group)
    queryset = apply_khong_khac_filter(queryset, "NgoaiNgu", ngoai_ngu_group)
    queryset = apply_khong_khac_filter(queryset, "TinHoc", tin_hoc_group)

    if tuoi_dang_nhom:
        today = timezone.localdate()

        def subtract_years(base_date, years):
            try:
                return base_date.replace(year=base_date.year - years)
            except ValueError:
                return base_date.replace(month=2, day=28, year=base_date.year - years)

        cutoff_10 = subtract_years(today, 10)
        cutoff_20 = subtract_years(today, 20)
        cutoff_30 = subtract_years(today, 30)
        cutoff_40 = subtract_years(today, 40)
        cutoff_50 = subtract_years(today, 50)

        if tuoi_dang_nhom == "lt10":
            queryset = queryset.filter(NgayVaoDang__isnull=False, NgayVaoDang__gt=cutoff_10)
        elif tuoi_dang_nhom == "gte10":
            queryset = queryset.filter(NgayVaoDang__isnull=False, NgayVaoDang__lte=cutoff_10)
        elif tuoi_dang_nhom == "gte20":
            queryset = queryset.filter(NgayVaoDang__isnull=False, NgayVaoDang__lte=cutoff_20)
        elif tuoi_dang_nhom == "gte30":
            queryset = queryset.filter(NgayVaoDang__isnull=False, NgayVaoDang__lte=cutoff_30)
        elif tuoi_dang_nhom == "gte40":
            queryset = queryset.filter(NgayVaoDang__isnull=False, NgayVaoDang__lte=cutoff_40)
        elif tuoi_dang_nhom == "gte50":
            queryset = queryset.filter(NgayVaoDang__isnull=False, NgayVaoDang__lte=cutoff_50)

    return queryset, {
        "q": keyword,
        "gioi_tinh": gioi_tinh,
        "dien": dien,
        "chi_bo": chi_bo,
        "trang_thai": trang_thai,
        "que_quan": que_quan,
        "dan_toc": dan_toc,
        "gdpt": gdpt,
        "gddh_group": gddh_group,
        "gdsdh_group": gdsdh_group,
        "hoc_ham_group": hoc_ham_group,
        "ngoai_ngu_group": ngoai_ngu_group,
        "tin_hoc_group": tin_hoc_group,
        "ly_luan_chinh_tri": ly_luan_chinh_tri,
        "tuoi_dang_nhom": tuoi_dang_nhom,
    }


def _get_dang_vien_list_with_metrics(reference_date=None, queryset=None):
    today = reference_date or timezone.localdate()
    current_year = today.year

    source_queryset = queryset if queryset is not None else _base_dang_vien_queryset()
    dangvien_list = list(source_queryset)

    ages = []
    party_ages = []
    for obj in dangvien_list:
        if obj.NgaySinh:
            ages.append(max(0, current_year - obj.NgaySinh.year))

        if obj.NgayVaoDang:
            obj.tuoi_dang = max(0, current_year - obj.NgayVaoDang.year)
            party_ages.append(obj.tuoi_dang)
        else:
            obj.tuoi_dang = "-"

        obj.canh_bao_qua_han_xet = obj.is_overdue_membership_review(reference_date=today)
        obj.han_xet_dang_vien = obj.get_review_deadline()

    def _sort_key(item):
        is_overdue_cho_xet = item.DienDangVien == "CHO_XET" and item.canh_bao_qua_han_xet
        deadline = item.han_xet_dang_vien or datetime.date.max
        return (0 if is_overdue_cho_xet else 1, deadline, item.HoTen or "")

    dangvien_list.sort(key=_sort_key)

    return dangvien_list, ages, party_ages


def _get_dang_vien_overview_context(reference_date=None):
    today = reference_date or timezone.localdate()
    current_year = today.year
    dangvien_list, ages, party_ages = _get_dang_vien_list_with_metrics(reference_date=today)

    total = len(dangvien_list)
    kinh_count = sum(1 for item in dangvien_list if (item.DanToc or "").strip().lower() == "kinh")
    active_members = sum(1 for item in dangvien_list if item.TrangThaiSinhHoat == "Đang sinh hoạt")
    inactive_members = total - active_members

    average_age = round(sum(ages) / len(ages)) if ages else 0
    average_party_age = round(sum(party_ages) / len(party_ages)) if party_ages else 0
    median_party_age = statistics.median(party_ages) if party_ages else 0

    min_age = min(ages) if ages else 0
    max_age = max(ages) if ages else 0

    def _format_int_vi(value):
        return f"{int(value):,}".replace(",", ".")

    def _format_decimal_vi(value):
        return f"{float(value):.1f}".replace(".", ",")

    def _format_percent_vi(value):
        rounded = round(value, 1)
        if float(rounded).is_integer():
            return f"{int(rounded)}"
        return _format_decimal_vi(rounded)

    def _build_donut_style(slices):
        if not slices:
            return "conic-gradient(#cbd5e1 0 100%)"
        start = 0.0
        parts = []
        for item in slices:
            end = start + item["percent"]
            parts.append(f"{item['color']} {start:.4f}% {end:.4f}%")
            start = end
        if start < 100:
            parts.append(f"#e2e8f0 {start:.4f}% 100%")
        return f"conic-gradient({', '.join(parts)})"

    def _build_filter_key(prefix, label):
        normalized = unicodedata.normalize("NFD", label or "")
        ascii_text = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
        clean_chars = []
        for ch in ascii_text.lower():
            if ch.isalnum():
                clean_chars.append(ch)
            elif clean_chars and clean_chars[-1] != "_":
                clean_chars.append("_")
        key_core = "".join(clean_chars).strip("_") or "khac"
        return f"{prefix}_{key_core}"

    def _build_top3_segments(raw_counts, colors, key_prefix):
        entries = [(label, count) for label, count in raw_counts.items() if count > 0]
        entries.sort(key=lambda item: item[1], reverse=True)
        if len(entries) > 3:
            entries = entries[:2] + [("Khác", sum(count for _, count in entries[2:]))]
        total_entries = sum(count for _, count in entries)
        segments = []
        for idx, (label, count) in enumerate(entries):
            percent = (count * 100 / total_entries) if total_entries else 0
            color = colors[idx % len(colors)]
            key_label = "khac" if label == "Khác" else label
            segments.append(
                {
                    "key": _build_filter_key(key_prefix, key_label),
                    "label": label,
                    "count": count,
                    "percent": percent,
                    "percent_text": _format_percent_vi(percent),
                    "color": color,
                }
            )
        return segments

    age_group_labels = ["Dưới 30", "30-40", "40-50", "50-60", "60-70", "Trên 70"]
    party_age_group_labels = ["Dưới 5 năm", "5-10 năm", "10-20 năm", "20-30 năm", "30-50 năm", "Trên 50 năm"]

    def _age_bucket_index(member):
        if not member.NgaySinh:
            return None
        age = max(0, current_year - member.NgaySinh.year)
        if age < 30: return 0
        if age < 40: return 1
        if age < 50: return 2
        if age < 60: return 3
        if age < 70: return 4
        return 5

    def _build_age_distribution_for_members(members):
        counts = [0] * len(age_group_labels)
        for member in members:
            idx = _age_bucket_index(member)
            if idx is not None:
                counts[idx] += 1
        max_count = max(counts) if counts else 0
        result = []
        for idx, (label, count) in enumerate(zip(age_group_labels, counts)):
            ratio = (count * 100 / max_count) if max_count else 0
            height_px = max(8, int(ratio * 2.0)) if count > 0 else 0
            result.append({"key": f"age_{idx}", "label": label, "count": count, "ratio": ratio, "height_px": height_px})
        return result

    def _party_age_bucket_index(member):
        if not member.NgayVaoDang:
            return None
        party_age = max(0, current_year - member.NgayVaoDang.year)
        if party_age < 5: return 0
        if party_age < 10: return 1
        if party_age < 20: return 2
        if party_age < 30: return 3
        if party_age < 50: return 4
        return 5

    def _build_party_age_distribution_for_members(members):
        counts = [0] * len(party_age_group_labels)
        for member in members:
            idx = _party_age_bucket_index(member)
            if idx is not None:
                counts[idx] += 1
        max_count = max(counts) if counts else 0
        result = []
        for idx, (label, count) in enumerate(zip(party_age_group_labels, counts)):
            ratio = (count * 100 / max_count) if max_count else 0
            height_px = max(8, int(ratio * 2.0)) if count > 0 else 0
            result.append({"key": f"party_age_{idx}", "label": label, "count": count, "ratio": ratio, "height_px": height_px})
        return result

    percent_kinh = round((kinh_count * 100 / total), 1) if total else 0
    percent_kinh_text = (
        f"{int(percent_kinh)}" if float(percent_kinh).is_integer() else _format_decimal_vi(percent_kinh)
    )

    age_range_text = "Không có dữ liệu"
    if ages:
        age_range_text = f"Min {min_age} - Max {max_age} tuổi"

    def _build_overview_payload(members):
        subset_total = len(members)
        male_count = sum(1 for item in members if item.GioiTinh == "Nam")
        female_count = sum(1 for item in members if item.GioiTinh == "Nữ")
        male_percent = (male_count * 100 / subset_total) if subset_total else 0
        female_percent = (female_count * 100 / subset_total) if subset_total else 0

        gender_segments = []
        if subset_total:
            gender_segments = [
                {"key": "gender_nam", "label": "Nam", "count": male_count, "percent": male_percent, "percent_text": _format_percent_vi(male_percent), "color": "#9f1239"},
                {"key": "gender_nu", "label": "Nữ", "count": female_count, "percent": female_percent, "percent_text": _format_percent_vi(female_percent), "color": "#e11d48"},
            ]

        dien_chinh_thuc_count = sum(1 for item in members if item.DienDangVien == "CHINH_THUC")
        dien_cho_xet_count = sum(1 for item in members if item.DienDangVien == "CHO_XET")
        dien_other_count = max(0, subset_total - dien_chinh_thuc_count - dien_cho_xet_count)
        dien_chinh_thuc_percent = (dien_chinh_thuc_count * 100 / subset_total) if subset_total else 0
        dien_cho_xet_percent = (dien_cho_xet_count * 100 / subset_total) if subset_total else 0
        dien_other_percent = (dien_other_count * 100 / subset_total) if subset_total else 0

        membership_segments = []
        if subset_total:
            membership_segments = [
                {"key": "dien_chinh_thuc", "label": "Chính thức", "count": dien_chinh_thuc_count, "percent": dien_chinh_thuc_percent, "percent_text": _format_percent_vi(dien_chinh_thuc_percent), "color": "#7f1d1d"},
                {"key": "dien_cho_xet", "label": "Chờ xét", "count": dien_cho_xet_count, "percent": dien_cho_xet_percent, "percent_text": _format_percent_vi(dien_cho_xet_percent), "color": "#b45309"},
            ]
            if dien_other_count:
                membership_segments.append({"key": "dien_khac", "label": "Khác", "count": dien_other_count, "percent": dien_other_percent, "percent_text": _format_percent_vi(dien_other_percent), "color": "#64748b"})

        active_subset = sum(1 for item in members if item.TrangThaiSinhHoat == "Đang sinh hoạt")
        inactive_subset = subset_total - active_subset
        tinh_trang_segments = []
        if subset_total:
            tinh_trang_segments = [
                {"key": "tinhtrang_dang_sinh_hoat", "label": "Đang sinh hoạt", "count": active_subset, "percent": (active_subset * 100 / subset_total), "percent_text": _format_percent_vi(active_subset * 100 / subset_total), "color": "#9f1239"},
                {"key": "tinhtrang_mien_sinh_hoat", "label": "Miễn sinh hoạt", "count": inactive_subset, "percent": (inactive_subset * 100 / subset_total), "percent_text": _format_percent_vi(inactive_subset * 100 / subset_total), "color": "#78716c"},
            ]

        nghe_nghiep_counts = {}
        ly_luan_counts = {}
        for member in members:
            nghe_nghiep = (member.NgheNghiep or "").strip()
            ly_luan = (member.LyLuanChinhTri or "").strip()
            if nghe_nghiep:
                nghe_nghiep_counts[nghe_nghiep] = nghe_nghiep_counts.get(nghe_nghiep, 0) + 1
            if ly_luan:
                ly_luan_counts[ly_luan] = ly_luan_counts.get(ly_luan, 0) + 1

        nghe_nghiep_segments = _build_top3_segments(nghe_nghiep_counts, ["#9f1239", "#b45309", "#475569"], "nghe")
        ly_luan_segments = _build_top3_segments(ly_luan_counts, ["#7f1d1d", "#9a3412", "#334155"], "lyluan")

        return {
            "total": subset_total,
            "male_count": male_count,
            "female_count": female_count,
            "male_percent_text": _format_percent_vi(male_percent),
            "female_percent_text": _format_percent_vi(female_percent),
            "dien_other_count": dien_other_count,
            "dien_chinh_thuc_percent_text": _format_percent_vi(dien_chinh_thuc_percent),
            "dien_cho_xet_percent_text": _format_percent_vi(dien_cho_xet_percent),
            "dien_other_percent_text": _format_percent_vi(dien_other_percent),
            "gender_segments": gender_segments,
            "membership_segments": membership_segments,
            "tinh_trang_segments": tinh_trang_segments,
            "nghe_nghiep_segments": nghe_nghiep_segments,
            "ly_luan_segments": ly_luan_segments,
            "gender_chart_style": _build_donut_style(gender_segments),
            "membership_chart_style": _build_donut_style(membership_segments),
            "nghe_nghiep_chart_style": _build_donut_style(nghe_nghiep_segments),
            "ly_luan_chart_style": _build_donut_style(ly_luan_segments),
            "tinh_trang_chart_style": _build_donut_style(tinh_trang_segments),
            "age_distribution": _build_age_distribution_for_members(members),
            "party_age_distribution": _build_party_age_distribution_for_members(members),
        }

    default_payload = _build_overview_payload(dangvien_list)
    overview_charts_by_key = {"__default__": default_payload}

    def _register_payload(filter_key, members):
        overview_charts_by_key[filter_key] = _build_overview_payload(members)

    _register_payload("gender_nam", [m for m in dangvien_list if m.GioiTinh == "Nam"])
    _register_payload("gender_nu", [m for m in dangvien_list if m.GioiTinh == "Nữ"])
    _register_payload("dien_chinh_thuc", [m for m in dangvien_list if m.DienDangVien == "CHINH_THUC"])
    _register_payload("dien_cho_xet", [m for m in dangvien_list if m.DienDangVien == "CHO_XET"])
    _register_payload("dien_khac", [m for m in dangvien_list if m.DienDangVien not in {"CHINH_THUC", "CHO_XET"}])
    _register_payload("tinhtrang_dang_sinh_hoat", [m for m in dangvien_list if m.TrangThaiSinhHoat == "Đang sinh hoạt"])
    _register_payload("tinhtrang_mien_sinh_hoat", [m for m in dangvien_list if m.TrangThaiSinhHoat != "Đang sinh hoạt"])

    nghe_values = sorted({(m.NgheNghiep or "").strip() for m in dangvien_list if (m.NgheNghiep or "").strip()})
    for nghe_label in nghe_values:
        nghe_key = _build_filter_key("nghe", nghe_label)
        _register_payload(nghe_key, [m for m in dangvien_list if (m.NgheNghiep or "").strip() == nghe_label])

    ly_luan_values = sorted({(m.LyLuanChinhTri or "").strip() for m in dangvien_list if (m.LyLuanChinhTri or "").strip()})
    for ly_luan_label in ly_luan_values:
        ly_key = _build_filter_key("lyluan", ly_luan_label)
        _register_payload(ly_key, [m for m in dangvien_list if (m.LyLuanChinhTri or "").strip() == ly_luan_label])

    top_nghe_labels = [item["label"] for item in default_payload["nghe_nghiep_segments"] if item["label"] != "Khác"]
    if any(item["label"] == "Khác" for item in default_payload["nghe_nghiep_segments"]):
        _register_payload(
            _build_filter_key("nghe", "khac"),
            [m for m in dangvien_list if (m.NgheNghiep or "").strip() and (m.NgheNghiep or "").strip() not in top_nghe_labels],
        )

    top_ly_luan_labels = [item["label"] for item in default_payload["ly_luan_segments"] if item["label"] != "Khác"]
    if any(item["label"] == "Khác" for item in default_payload["ly_luan_segments"]):
        _register_payload(
            _build_filter_key("lyluan", "khac"),
            [m for m in dangvien_list if (m.LyLuanChinhTri or "").strip() and (m.LyLuanChinhTri or "").strip() not in top_ly_luan_labels],
        )

    for age_idx, _ in enumerate(age_group_labels):
        _register_payload(f"age_{age_idx}", [m for m in dangvien_list if _age_bucket_index(m) == age_idx])

    for party_age_idx, _ in enumerate(party_age_group_labels):
        _register_payload(f"party_age_{party_age_idx}", [m for m in dangvien_list if _party_age_bucket_index(m) == party_age_idx])

    male_count = default_payload["male_count"]
    female_count = default_payload["female_count"]
    male_percent_text = default_payload["male_percent_text"]
    female_percent_text = default_payload["female_percent_text"]
    dien_other_count = default_payload["dien_other_count"]
    dien_chinh_thuc_percent_text = default_payload["dien_chinh_thuc_percent_text"]
    dien_cho_xet_percent_text = default_payload["dien_cho_xet_percent_text"]
    dien_other_percent_text = default_payload["dien_other_percent_text"]
    gender_segments = default_payload["gender_segments"]
    membership_segments = default_payload["membership_segments"]
    nghe_nghiep_segments = default_payload["nghe_nghiep_segments"]
    ly_luan_segments = default_payload["ly_luan_segments"]
    tinh_trang_segments = default_payload["tinh_trang_segments"]
    gender_chart_style = default_payload["gender_chart_style"]
    membership_chart_style = default_payload["membership_chart_style"]
    age_distribution = default_payload["age_distribution"]
    party_age_distribution = default_payload["party_age_distribution"]
    age_distribution_by_key = {key: payload["age_distribution"] for key, payload in overview_charts_by_key.items()}

    return {
        "total": total,
        "total_display": _format_int_vi(total),
        "male": male_count,
        "female": female_count,
        "active_members": active_members,
        "inactive_members": inactive_members,
        "average_age": average_age,
        "average_party_age": average_party_age,
        "average_age_display": _format_int_vi(average_age),
        "average_party_age_display": _format_int_vi(average_party_age),
        "median_party_age_display": _format_int_vi(median_party_age),
        "kinh_percent_text": percent_kinh_text,
        "age_range_text": age_range_text,
        "male_percent_text": male_percent_text,
        "female_percent_text": female_percent_text,
        "dien_chinh_thuc_percent_text": dien_chinh_thuc_percent_text,
        "dien_cho_xet_percent_text": dien_cho_xet_percent_text,
        "dien_other_percent_text": dien_other_percent_text,
        "dien_other_count": dien_other_count,
        "gender_segments": gender_segments,
        "membership_segments": membership_segments,
        "gender_chart_style": gender_chart_style,
        "membership_chart_style": membership_chart_style,
        "age_distribution": age_distribution,
        "party_age_distribution": party_age_distribution,
        "age_distribution_by_key": age_distribution_by_key,
        "overview_charts_by_key": overview_charts_by_key,
        "nghe_nghiep_segments": nghe_nghiep_segments,
        "nghe_nghiep_chart_style": _build_donut_style(nghe_nghiep_segments),
        "ly_luan_segments": ly_luan_segments,
        "ly_luan_chart_style": _build_donut_style(ly_luan_segments),
        "tinh_trang_segments": tinh_trang_segments,
        "tinh_trang_chart_style": _build_donut_style(tinh_trang_segments),
    }


def _export_dang_vien_csv(dangvien_list):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="dang_vien_data.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(["CCCD","Mã ĐV","Họ tên","Bí danh","Giới tính","Ngày sinh","Quê quán","Dân tộc","Nơi thường trú","Nơi tạm trú","Nghề nghiệp","GDPT","GDNN","GDDH","GDSĐH","Học hàm","Lý luận chính trị","Ngoại ngữ","Tin học","Tiếng DTTS","Ngày vào Đảng","Ngày chính thức","Tuổi Đảng","Đảng bộ","Chi bộ","Trạng thái","Diện Đảng viên","Số điện thoại","Ghi chú","Số Huy hiệu","Huy hiệu cao nhất"])
    for item in dangvien_list:
        writer.writerow([
            item.SoCCCD,
            item.MaDangVien,
            item.HoTen,
            item.BiDanh,
            item.GioiTinh,
            item.NgaySinh.strftime("%d/%m/%Y") if item.NgaySinh else "",
            item.QueQuan,
            item.DanToc,
            item.NoiThuongTru,
            item.NoiTamTru,
            item.NgheNghiep,
            item.GDPT,
            item.GDNN,
            item.GDDH,
            item.GDSĐH,
            item.HocHam,
            item.LyLuanChinhTri,
            item.NgoaiNgu,
            item.TinHoc,
            item.TiengDTTS,
            item.NgayVaoDang.strftime("%d/%m/%Y") if item.NgayVaoDang else "",
            item.NgayChinhThuc.strftime("%d/%m/%Y") if item.NgayChinhThuc else "",
            item.tuoi_dang,
            item.DangBoID.TenDangBo if item.DangBoID else "",
            item.ChiBoID.TenChiBo if item.ChiBoID else "",
            item.TrangThaiSinhHoat,
            item.DienDangVien,
            item.SoDienThoai,
            item.GhiChu,
            item.so_huy_hieu,
            item.HuyHieuCaoNhat or "",
        ])

    return response


def _export_dang_vien_excel(dangvien_list):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("Thiếu thư viện openpyxl để xuất Excel. Vui lòng cài đặt openpyxl.", status=500, content_type="text/plain; charset=utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "DangVien"
    headers = ["CCCD","Mã ĐV","Họ tên","Bí danh","Giới tính","Ngày sinh","Quê quán","Dân tộc","Nơi thường trú","Nơi tạm trú","Nghề nghiệp","GDPT","GDNN","GDDH","GDSĐH","Học hàm","Lý luận chính trị","Ngoại ngữ","Tin học","Tiếng DTTS","Ngày vào Đảng","Ngày chính thức","Tuổi Đảng","Đảng bộ","Chi bộ","Trạng thái","Diện Đảng viên","Số điện thoại","Ghi chú","Số Huy hiệu","Huy hiệu cao nhất"]
    worksheet.append(headers)
    for item in dangvien_list:
        worksheet.append([
            item.SoCCCD,
            item.MaDangVien,
            item.HoTen,
            item.BiDanh,
            item.GioiTinh,
            item.NgaySinh.strftime("%d/%m/%Y") if item.NgaySinh else "",
            item.QueQuan,
            item.DanToc,
            item.NoiThuongTru,
            item.NoiTamTru,
            item.NgheNghiep,
            item.GDPT,
            item.GDNN,
            item.GDDH,
            item.GDSĐH,
            item.HocHam,
            item.LyLuanChinhTri,
            item.NgoaiNgu,
            item.TinHoc,
            item.TiengDTTS,
            item.NgayVaoDang.strftime("%d/%m/%Y") if item.NgayVaoDang else "",
            item.NgayChinhThuc.strftime("%d/%m/%Y") if item.NgayChinhThuc else "",
            item.tuoi_dang,
            item.DangBoID.TenDangBo if item.DangBoID else "",
            item.ChiBoID.TenChiBo if item.ChiBoID else "",
            item.TrangThaiSinhHoat,
            item.DienDangVien,
            item.SoDienThoai,
            item.GhiChu,
            item.so_huy_hieu,
            item.HuyHieuCaoNhat or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="dang_vien_data.xlsx"'
    workbook.save(response)
    return response


def _get_dang_vien_structure_context(reference_date=None):
    today = reference_date or timezone.localdate()
    members = list(DangVien.objects.select_related("ChiBoID").all())
    total = len(members)

    period_ranges = [
        ("Trước 1975", None, 1974),
        ("1975 - 1986", 1975, 1986),
        ("1986 - 2000", 1987, 2000),
        ("2000 - 2010", 2001, 2010),
        ("2010 - nay", 2011, None),
    ]

    period_cards = []
    for label, start_year, end_year in period_ranges:
        count = 0
        for member in members:
            if not member.NgayVaoDang:
                continue
            year = member.NgayVaoDang.year
            if start_year is None and end_year is not None:
                if year <= end_year:
                    count += 1
            elif end_year is None and start_year is not None:
                if year >= start_year:
                    count += 1
            elif start_year is not None and end_year is not None and start_year <= year <= end_year:
                count += 1
        percent = round((count * 100 / total), 1) if total else 0
        period_cards.append({"label": label, "count": count, "percent": percent})

    period_colors = ["#9f1239", "#be123c", "#b45309", "#9a3412", "#7f1d1d"]
    period_segments = []
    for idx, item in enumerate(period_cards):
        period_segments.append({"label": item["label"], "count": item["count"], "percent": item["percent"], "percent_text": str(item["percent"]).replace(".", ","), "color": period_colors[idx % len(period_colors)]})

    if total:
        start_percent = 0.0
        donut_parts = []
        for item in period_segments:
            end_percent = start_percent + item["percent"]
            donut_parts.append(f"{item['color']} {start_percent:.4f}% {end_percent:.4f}%")
            start_percent = end_percent
        if start_percent < 100:
            donut_parts.append(f"#e2e8f0 {start_percent:.4f}% 100%")
        period_donut_style = f"conic-gradient({', '.join(donut_parts)})"
    else:
        period_donut_style = "conic-gradient(#cbd5e1 0 100%)"

    month_counts = {month: 0 for month in range(1, 13)}
    for member in members:
        if member.NgayVaoDang:
            month_counts[member.NgayVaoDang.month] += 1

    top_month_pairs = sorted(month_counts.items(), key=lambda item: (-item[1], item[0]))[:5]
    top_month_pairs.sort(key=lambda item: item[0])
    max_top_month = max((count for _, count in top_month_pairs), default=0)
    monthly_admission_top = []
    for month, count in top_month_pairs:
        ratio = (count * 100 / max_top_month) if max_top_month else 0
        monthly_admission_top.append({"label": f"Tháng {month}", "value": count, "ratio": ratio, "height_px": max(6, int(ratio * 1.9)) if count else 3})

    yearly_values = [member.NgayVaoDang.year for member in members if member.NgayVaoDang]
    start_year = 1944
    end_year = max(yearly_values) if yearly_values else today.year
    yearly_counts = {year: 0 for year in range(start_year, end_year + 1)}
    for member in members:
        if member.NgayVaoDang and start_year <= member.NgayVaoDang.year <= end_year:
            yearly_counts[member.NgayVaoDang.year] += 1

    max_yearly = max(yearly_counts.values()) if yearly_counts else 0
    special_year = None
    special_count = 0
    if max_yearly > 0:
        special_year = min((year for year, count in yearly_counts.items() if count == max_yearly), default=None)
        special_count = max_yearly

    def _year_color(year):
        if year <= 1974: return "#d65a2b"
        if year <= 1985: return "#2b6cb0"
        if year <= 1999: return "#1f9d79"
        if year <= 2009: return "#5b50c8"
        return "#eb9e26"

    yearly_admission_bars = []
    for year in range(start_year, end_year + 1):
        value = yearly_counts[year]
        ratio = (value * 100 / max_yearly) if max_yearly else 0
        height_px = max(2, int(ratio * 3.4)) if value > 0 else 1
        yearly_admission_bars.append({"year": year, "value": value, "height_px": height_px, "color": _year_color(year), "is_special": special_year is not None and year == special_year, "show_label": (year % 5 == 0)})

    return {
        "total": total,
        "period_cards": period_cards,
        "period_segments": period_segments,
        "period_donut_style": period_donut_style,
        "year_chart_start": start_year,
        "year_chart_end": end_year,
        "yearly_admission_bars": yearly_admission_bars,
        "yearly_special_year": special_year,
        "yearly_special_count": special_count,
        "monthly_admission_top": monthly_admission_top,
    }


def dangvien(request):
    _sync_dang_vien_membership_statuses()
    context = _get_dang_vien_overview_context()
    return render(request, "qldv/dangvien.html", context)


def dangvien_structure(request):
    _sync_dang_vien_membership_statuses()
    context = _get_dang_vien_structure_context()
    return render(request, "qldv/dangvien_structure.html", context)


def dangvien_data(request):
    _sync_dang_vien_membership_statuses()
    queryset = _base_dang_vien_queryset()
    filtered_queryset, active_filters = _apply_dang_vien_filters(queryset, request.GET)
    dangvien_list, _, _ = _get_dang_vien_list_with_metrics(reference_date=timezone.localdate(), queryset=filtered_queryset)

    download_format = (request.GET.get("download") or "").strip().lower()
    if download_format == "csv":
        return _export_dang_vien_csv(dangvien_list)
    if download_format == "excel":
        return _export_dang_vien_excel(dangvien_list)

    context = {
        "dangvien_list": dangvien_list,
        "total_filtered": len(dangvien_list),
        "filters": active_filters,
        "current_full_path": request.get_full_path(),
        "overdue_cho_xet_members": [item for item in dangvien_list if item.DienDangVien == "CHO_XET" and item.canh_bao_qua_han_xet],
        "chi_bo_options": ChiBo.objects.order_by("TenChiBo"),
        "gioi_tinh_choices": [("Nam", "Nam"), ("Nữ", "Nữ")],
        "dien_dang_vien_choices": DangVien.DIEN_DANG_VIEN_CHOICES,
        "trang_thai_choices": DangVien.TRANG_THAI_SINH_HOAT_CHOICES,
        "que_quan_options": [item[0] for item in PROVINCE_CITY_CHOICES if item[0]],
        "dan_toc_options": [item[0] for item in ETHNIC_GROUPS if item[0]],
        "gdpt_options": [f"{i}/12" for i in range(1, 13)],
        "ly_luan_chinh_tri_options": ["Không", "Sơ cấp", "Trung cấp", "Cao cấp", "Cử nhân"],
        "khong_khac_options": [("KHONG", "Không"), ("KHAC", "Khác")],
        "tuoi_dang_groups": [("lt10", "Dưới 10 năm"), ("gte10", "10 năm trở lên"), ("gte20", "20 năm trở lên"), ("gte30", "30 năm trở lên"), ("gte40", "40 năm trở lên"), ("gte50", "50 năm trở lên")],
    }
    return render(request, "qldv/dangvien_data.html", context)


def dangvien_update_dien(request, dangvien_id):
    if request.method != "POST":
        return redirect("dangvien_data")

    dangvien_item = get_object_or_404(DangVien, pk=dangvien_id)
    new_dien = (request.POST.get("dien_dang_vien") or "").strip()
    valid_dien_values = {value for value, _ in DangVien.DIEN_DANG_VIEN_CHOICES}
    next_path = (request.POST.get("next") or "").strip()

    if new_dien not in valid_dien_values:
        messages.error(request, "Diện Đảng viên không hợp lệ.")
    else:
        dangvien_item.DienDangVien = new_dien
        dangvien_item.apply_membership_rules(reference_date=timezone.localdate())
        dangvien_item.save(update_fields=["DienDangVien", "NgayChinhThuc", "NgayVaoDang"])
        messages.success(request, "Đã cập nhật Diện Đảng viên.")

    if next_path.startswith("/"):
        return redirect(next_path)
    return redirect("dangvien_data")


@login_required(login_url='login')
def dangvien_add(request):
    default_dang_bo = _get_single_dang_bo()

    if request.method == "POST" and request.POST.get("form_type") == "import":
        single_form = DangVienForm(dang_bo=default_dang_bo)
        import_form = DangVienImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            try:
                created, errors = _import_dang_vien_data(import_form.cleaned_data["import_file"], default_dang_bo=default_dang_bo)
            except Exception as exc:
                messages.error(request, f"Không thể đọc file import: {exc}")
            else:
                if created:
                    messages.success(request, f"Import thành công {created} bản ghi.")
                if errors:
                    messages.warning(request, "\n".join(errors[:20]))
                if created and not errors:
                    return redirect("dangvien_add")
    elif request.method == "POST":
        single_form = DangVienForm(request.POST, dang_bo=default_dang_bo)
        import_form = DangVienImportForm()
        if single_form.is_valid():
            single_form.save()
            messages.success(request, "Đã lưu thông tin Đảng viên.")
            return redirect("dangvien_add")
        messages.error(request, "Không thể lưu Đảng viên. Vui lòng kiểm tra các trường đang báo lỗi.")
    else:
        single_form = DangVienForm(dang_bo=default_dang_bo)
        import_form = DangVienImportForm()

    return render(request, "qldv/dangvien_add.html", {"single_form": single_form, "import_form": import_form})


def dangvien_edit(request, dangvien_id):
    dangvien_item = get_object_or_404(DangVien, pk=dangvien_id)
    dang_bo = dangvien_item.DangBoID

    if request.method == "POST":
        form = DangVienForm(request.POST, instance=dangvien_item, dang_bo=dang_bo)
        if form.is_valid():
            form.save()
            messages.success(request, "Đã cập nhật thông tin Đảng viên.")
            return redirect("dangvien")
    else:
        form = DangVienForm(instance=dangvien_item, dang_bo=dang_bo)

    return render(request, "qldv/dangvien_edit.html", {"form": form, "dangvien_item": dangvien_item})


def _base_huy_hieu_queryset():
    return HuyHieuDang.objects.select_related("DangVienID", "DangVienID__ChiBoID").order_by("-NgayTrao")


def _apply_huy_hieu_filters(queryset, query_params):
    q = (query_params.get("q") or "").strip()
    year = (query_params.get("year") or "").strip()
    badge_year = (query_params.get("badge_year") or "").strip()
    trang_thai = (query_params.get("trang_thai") or "").strip()

    if q:
        queryset = queryset.filter(Q(DangVienID__HoTen__icontains=q) | Q(DangVienID__MaDangVien__icontains=q) | Q(DangVienID__SoCCCD__icontains=q))
    if year and year.isdigit():
        queryset = queryset.filter(NgayTrao__year=int(year))
    if badge_year:
        queryset = queryset.filter(LoaiHuyHieu__icontains=f"{badge_year} năm")
    if trang_thai:
        queryset = queryset.filter(TrangThai=trang_thai)

    return queryset, {"q": q, "year": year, "badge_year": badge_year, "trang_thai": trang_thai}


def _export_huy_hieu_csv(huyhieu_list):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="huy_hieu_data.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(["Họ tên","Mã ĐV","CCCD","Loại huy hiệu","Ngày đủ điều kiện","Đợt trao tặng","Trạng thái","Số quyết định","Ngày trao","Ghi chú"])
    for item in huyhieu_list:
        writer.writerow([item.DangVienID.HoTen,item.DangVienID.MaDangVien,item.DangVienID.SoCCCD,item.LoaiHuyHieu,item.NgayDuDieuKien.strftime("%d/%m/%Y") if item.NgayDuDieuKien else "",item.DotTraoTang,item.TrangThai,item.SoQuyetDinh,item.NgayTrao.strftime("%d/%m/%Y") if item.NgayTrao else "",item.GhiChu])
    return response


def _export_huy_hieu_excel(huyhieu_list):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("Thiếu thư viện openpyxl để xuất Excel. Vui lòng cài đặt openpyxl.", status=500, content_type="text/plain; charset=utf-8")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "HuyHieu"
    worksheet.append(["Họ tên","Mã ĐV","CCCD","Loại huy hiệu","Ngày đủ điều kiện","Đợt trao tặng","Trạng thái","Số quyết định","Ngày trao","Ghi chú"])
    for item in huyhieu_list:
        worksheet.append([item.DangVienID.HoTen,item.DangVienID.MaDangVien,item.DangVienID.SoCCCD,item.LoaiHuyHieu,item.NgayDuDieuKien.strftime("%d/%m/%Y") if item.NgayDuDieuKien else "",item.DotTraoTang,item.TrangThai,item.SoQuyetDinh,item.NgayTrao.strftime("%d/%m/%Y") if item.NgayTrao else "",item.GhiChu])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="huy_hieu_data.xlsx"'
    workbook.save(response)
    return response


def dangvien_delete(request, dangvien_id):
    if request.method != "POST":
        return redirect("dangvien")
    dangvien_item = get_object_or_404(DangVien, pk=dangvien_id)
    dangvien_item.delete()
    messages.success(request, "Đã xóa Đảng viên thành công.")
    return redirect("dangvien")


def huyhieu(request):
    today = timezone.localdate()
    selected_year = request.GET.get("year", str(today.year))
    try:
        selected_year_int = int(selected_year)
    except ValueError:
        selected_year_int = today.year

    ceremony_dates = get_ceremony_dates(selected_year_int)
    selected_ceremony_key = request.GET.get("ceremony", "all")
    if selected_ceremony_key == "all":
        selected_ceremony = {"key": "all", "name": "Tất cả các đợt", "short_name": "cả năm", "date": datetime.date(selected_year_int, 12, 31)}
    else:
        selected_ceremony = next((item for item in ceremony_dates if item["key"] == selected_ceremony_key), ceremony_dates[0])

    review_date = selected_ceremony["date"]
    review_date_label = f"Tất cả các đợt năm {selected_year_int}" if selected_ceremony_key == "all" else review_date.strftime("%d/%m/%Y")

    badge_year_raw = request.GET.get("badge_year", "")
    try:
        selected_badge_year = int(badge_year_raw) if badge_year_raw else None
    except ValueError:
        selected_badge_year = None
    if selected_badge_year and selected_badge_year not in BADGE_MILESTONES:
        selected_badge_year = None

    year_options = list(range(today.year, 1929, -1))

    def build_context():
        huyhieu_list = HuyHieuDang.objects.select_related("DangVienID").order_by("-NgayTrao")
        dangvien_list = DangVien.objects.select_related("ChiBoID", "DangBoID").order_by("HoTen")
        eligible_members = get_eligible_members(dangvien_list, review_date, selected_badge_year, list(huyhieu_list))
        return {
            "total_huyhieu": huyhieu_list.count(),
            "review_date": review_date,
            "selected_year": selected_year_int,
            "selected_ceremony_key": selected_ceremony["key"],
            "review_date_label": review_date_label,
            "selected_badge_year": selected_badge_year,
            "ceremony_dates": ceremony_dates,
            "badge_milestones": BADGE_MILESTONES,
            "eligible_members": eligible_members,
            "total_eligible": len(eligible_members),
            "year_options": year_options,
        }

    if request.method == "POST":
        dang_vien_id = request.POST.get("dang_vien_id")
        eligible_years_raw = request.POST.get("eligible_for_years")
        eligible_date_raw = request.POST.get("eligible_date")
        so_quyet_dinh = request.POST.get("so_quyet_dinh", "").strip()
        ngay_trao_raw = request.POST.get("ngay_trao", "").strip()
        ghi_chu = request.POST.get("ghi_chu", "").strip()
        la_truy_tang = request.POST.get("la_truy_tang") == "on"

        if not dang_vien_id or not eligible_years_raw or not so_quyet_dinh or not ngay_trao_raw:
            messages.error(request, "Vui lòng nhập đầy đủ thông tin bắt buộc để lưu huy hiệu.")
            return redirect(request.get_full_path())

        try:
            eligible_for_years = int(eligible_years_raw)
            eligible_date = datetime.datetime.strptime(eligible_date_raw, "%Y-%m-%d").date()
            ngay_trao = datetime.datetime.strptime(ngay_trao_raw, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            messages.error(request, "Dữ liệu ngày hoặc mốc huy hiệu không hợp lệ.")
            return redirect(request.get_full_path())

        so_qd_valid, so_qd_error = validate_badge_decision_number(so_quyet_dinh, eligible_for_years)
        if not so_qd_valid:
            messages.error(request, so_qd_error)
            return redirect(request.get_full_path())

        dang_vien = get_object_or_404(DangVien, pk=dang_vien_id)
        existing_badge = HuyHieuDang.objects.filter(DangVienID=dang_vien, LoaiHuyHieu__icontains=f"{eligible_for_years} năm").exists()
        if existing_badge:
            messages.warning(request, f"{dang_vien.HoTen} đã có huy hiệu mốc {eligible_for_years} năm.")
            return redirect(request.get_full_path())

        dot_trao_tang = selected_ceremony["name"]
        if selected_ceremony_key == "all":
            dot_trao_tang = f"Tất cả các đợt năm {selected_year_int}"

        trang_thai = "Truy tặng" if la_truy_tang else "Đã trao"
        if la_truy_tang:
            ghi_chu = f"[Truy tặng] {ghi_chu}".strip()

        HuyHieuDang.objects.create(
            DangVienID=dang_vien,
            LoaiHuyHieu=f"Huy hiệu {eligible_for_years} năm tuổi Đảng",
            NgayDuDieuKien=eligible_date,
            DotTraoTang=dot_trao_tang,
            TrangThai=trang_thai,
            SoQuyetDinh=so_quyet_dinh,
            NgayTrao=ngay_trao,
            GhiChu=ghi_chu or None,
        )
        messages.success(request, f"Đã lưu huy hiệu cho {dang_vien.HoTen} ({trang_thai.lower()}).")
        return redirect(request.get_full_path())

    return render(request, "qldv/huyhieu.html", build_context())


def huyhieu_add(request):
    if request.method == "POST":
        action_type = request.POST.get("action_type")
        if action_type == "legacy_single":
            form = HuyHieuLegacyForm(request.POST)
            if form.is_valid():
                data = form.cleaned_data
                dang_vien = _resolve_dang_vien_for_huy_hieu(ma_dang_vien=data.get("MaDangVien"), so_cccd=data.get("SoCCCD"))
                if not dang_vien:
                    messages.error(request, "Không tìm thấy Đảng viên.")
                else:
                    duplicated = HuyHieuDang.objects.filter(DangVienID=dang_vien, LoaiHuyHieu=data["LoaiHuyHieu"], SoQuyetDinh=data["SoQuyetDinh"]).exists()
                    if duplicated:
                        messages.warning(request, "Hồ sơ huy hiệu này đã tồn tại.")
                    else:
                        HuyHieuDang.objects.create(DangVienID=dang_vien, LoaiHuyHieu=data["LoaiHuyHieu"], NgayDuDieuKien=data["NgayDuDieuKien"], DotTraoTang=data["DotTraoTang"], TrangThai=data["TrangThai"], SoQuyetDinh=data["SoQuyetDinh"], NgayTrao=data["NgayTrao"], GhiChu=(data.get("GhiChu") or "").strip() or None)
                        messages.success(request, f"Đã lưu hồ sơ cho {dang_vien.HoTen}.")
                        return redirect("huyhieu_add")
            else:
                messages.error(request, "Dữ liệu không hợp lệ.")
        elif action_type == "legacy_import":
            form = HuyHieuImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    created, errors = _import_huy_hieu_data(form.cleaned_data["import_file"])
                    if created:
                        messages.success(request, f"Đã import {created} hồ sơ.")
                    if errors:
                        messages.warning(request, "\n".join(errors[:10]))
                    if created:
                        return redirect("huyhieu_add")
                except Exception as exc:
                    messages.error(request, f"Lỗi import: {exc}")

    return render(request, "qldv/huyhieu_add.html", {"legacy_form": HuyHieuLegacyForm(), "import_form": HuyHieuImportForm()})


def huyhieu_data(request):
    queryset = _base_huy_hieu_queryset()
    filtered_queryset, active_filters = _apply_huy_hieu_filters(queryset, request.GET)
    huyhieu_list = list(filtered_queryset)

    download_format = (request.GET.get("download") or "").strip().lower()
    if download_format == "csv":
        return _export_huy_hieu_csv(huyhieu_list)
    if download_format == "excel":
        return _export_huy_hieu_excel(huyhieu_list)

    context = {
        "huyhieu_list": huyhieu_list,
        "total_filtered": len(huyhieu_list),
        "filters": active_filters,
        "year_options": range(timezone.localdate().year, 1929, -1),
        "badge_milestones": BADGE_MILESTONES,
        "trang_thai_choices": ["Đã trao", "Truy tặng"],
        "today": timezone.localdate(),
    }
    return render(request, "qldv/huyhieu_data.html", context)


def huyhieu_edit(request, huyhieu_id):
    huyhieu_item = get_object_or_404(HuyHieuDang, pk=huyhieu_id)

    if request.method == "POST":
        form = HuyHieuEditForm(request.POST, instance=huyhieu_item)
        if form.is_valid():
            form.save()
            messages.success(request, "Đã cập nhật hồ sơ Huy hiệu Đảng.")
            return redirect("huyhieu")
        messages.error(request, "Không thể cập nhật hồ sơ. Vui lòng kiểm tra các trường đang báo lỗi.")
    else:
        form = HuyHieuEditForm(instance=huyhieu_item)

    return render(request, "qldv/huyhieu_edit.html", {"form": form, "huyhieu_item": huyhieu_item})


def huyhieu_delete(request, huyhieu_id):
    if request.method != "POST":
        return redirect("huyhieu")
    huyhieu_item = get_object_or_404(HuyHieuDang, pk=huyhieu_id)
    ho_ten = huyhieu_item.DangVienID.HoTen
    loai_huy_hieu = huyhieu_item.LoaiHuyHieu
    huyhieu_item.delete()
    messages.success(request, f"Đã xóa hồ sơ {loai_huy_hieu} của {ho_ten}.")
    return redirect("huyhieu")


def user(request):
    return render(request, "qldv/user.html")


# ==================== Authentication Views ====================

def login_view(request):
    from django.contrib.auth import authenticate, login
    from .forms import LoginForm

    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            identifier = form.cleaned_data.get("identifier")
            password = form.cleaned_data.get("password")
            user = authenticate(request, username=identifier, password=password)
            if user is not None:
                login(request, user)
                # Kiểm tra nếu người dùng cần đổi mật khẩu
                if hasattr(user, 'profile') and user.profile.must_change_password:
                    messages.info(request, "Đây là lần đăng nhập đầu tiên, vui lòng đổi mật khẩu để tiếp tục.")
                    return redirect("change_password")
                return redirect("home")
            else:
                messages.error(request, "Tên đăng nhập hoặc mật khẩu không chính xác.")
    else:
        form = LoginForm()

    return render(request, "qldv/login.html", {"form": form})


def signup_view(request):
    """Đã tắt chức năng đăng ký tài khoản"""
    return redirect("login")


@login_required(login_url='login')
def change_password_view(request):
    from django.contrib.auth import update_session_auth_hash
    from django.contrib.auth.forms import PasswordChangeForm
    
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Cập nhật session để không bị logout sau khi đổi mật khẩu
            update_session_auth_hash(request, user)
            
            # Cập nhật trạng thái đã đổi mật khẩu
            if hasattr(user, 'profile'):
                user.profile.must_change_password = False
                user.profile.save()
                
            messages.success(request, "Đổi mật khẩu thành công.")
            return redirect("home")
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, "qldv/change_password.html", {"form": form})


def logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    messages.success(request, "Bạn đã đăng xuất thành công.")
    return redirect("home")